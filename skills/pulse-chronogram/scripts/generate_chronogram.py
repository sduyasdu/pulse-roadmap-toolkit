"""
Generate a weekly chronogram (Gantt-style) spreadsheet from a Pulse task
list, always rendered in the Yasdu brand style.

Usage: python3 generate_chronogram.py <input.json> <output.xlsx> <skill_dir>

<skill_dir> is the path to this skill's root (the folder containing this
script's parent), used to locate the assets/yasdu/ logo.

See ../references/input_schema.md for the full shape of input.json.
"""
import sys
import os
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

if len(sys.argv) not in (3, 4):
    print("Usage: python3 generate_chronogram.py <input.json> <output.xlsx> [skill_dir]")
    sys.exit(1)

INPUT_PATH, OUTPUT_PATH = sys.argv[1], sys.argv[2]
SKILL_DIR = sys.argv[3] if len(sys.argv) == 4 else os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

with open(INPUT_PATH, encoding="utf-8") as f:
    data = json.load(f)

meta = data["meta"]
tasks = data["tasks"]  # [{epic, title, start, end}, ...]
epic_order = data.get("epicOrder") or sorted({t["epic"] for t in tasks})
n_weeks = int(data.get("weeks", 8))
start_anchor = datetime.date.fromisoformat(data["startDate"])
month_names = data.get("monthNames") or {
    "1": "Ene", "2": "Feb", "3": "Mar", "4": "Abr", "5": "May", "6": "Jun",
    "7": "Jul", "8": "Ago", "9": "Sep", "10": "Oct", "11": "Nov", "12": "Dic",
}

# ---------------- Style: always Yasdu brand ----------------
S = {
    "font": "Calibri",  # fallback for Inter; header cells use bold for emphasis
    "header_fill": "D85A28",       # Yasdu brand orange
    "month_fill": "123359",        # Yasdu brand navy
    "epic_fill": "F7E8DA",         # Yasdu accent bg
    "epic_text": "123359",
    "active_fill": "D85A28",
    "title_color": "123359",
    "border": "E2DFD9",
    "alt_row": "F4F2EC",
    "text": "1F2330",
    "muted": "6E7180",
    "header_text": "FDFDFD",
    "logo": os.path.join(SKILL_DIR, "assets", "yasdu", "logo-horizontal-light.png"),
    "logo_wh_px": (1540, 443),
}
FONT_NAME = S["font"]

def pdate(s):
    return datetime.date.fromisoformat(s)

WEEK_START = start_anchor - datetime.timedelta(days=start_anchor.weekday())
weeks = []
for i in range(n_weeks):
    ws_ = WEEK_START + datetime.timedelta(weeks=i)
    we_ = ws_ + datetime.timedelta(days=6)
    weeks.append((ws_, we_))

thin = Side(style="thin", color=S["border"])
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = meta.get("sheetName", "Cronograma")

TASK_COL_WIDTH = 28
WEEK_COL_WIDTH = 9.5

ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = TASK_COL_WIDTH
for i in range(n_weeks):
    col = get_column_letter(3 + i)
    ws.column_dimensions[col].width = WEEK_COL_WIDTH

TITLE_ROW = 1
SUBTITLE_ROW = 2
LOGO_ROW_SPAN = 0

# ---------------- Logo (if any) ----------------
if S["logo"] and os.path.exists(S["logo"]):
    img = XLImage(S["logo"])
    px_w, px_h = S["logo_wh_px"]
    target_h_px = 32
    scale = target_h_px / px_h
    img.width = px_w * scale
    img.height = target_h_px
    ws.add_image(img, "A1")
    TITLE_ROW = 4
    SUBTITLE_ROW = 5
    ws.row_dimensions[1].height = 28

# ---------------- Title rows ----------------
ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=2 + n_weeks)
c = ws.cell(row=TITLE_ROW, column=1, value=meta["title"])
c.font = Font(name=FONT_NAME, size=18, bold=True, color=S["title_color"])

subtitle = meta.get("subtitle") or ""
if "{start}" in subtitle or "{end}" in subtitle or "{n}" in subtitle:
    subtitle = subtitle.format(
        n=n_weeks,
        start=WEEK_START.strftime("%d/%m/%Y"),
        end=weeks[-1][1].strftime("%d/%m/%Y"),
    )
ws.merge_cells(start_row=SUBTITLE_ROW, start_column=1, end_row=SUBTITLE_ROW, end_column=2 + n_weeks)
c = ws.cell(row=SUBTITLE_ROW, column=1, value=subtitle)
c.font = Font(name=FONT_NAME, size=11, italic=True, color=S["muted"])

HEADER_ROW_MONTH = SUBTITLE_ROW + 2
HEADER_ROW_WEEK = HEADER_ROW_MONTH + 1
FIRST_DATA_ROW = HEADER_ROW_WEEK + 1

# ---------------- Month header (merged across weeks belonging to it) ----------------
week_months = [(w[0].year, w[0].month) for w in weeks]
i = 0
while i < n_weeks:
    j = i
    while j + 1 < n_weeks and week_months[j + 1] == week_months[i]:
        j += 1
    start_col, end_col = 3 + i, 3 + j
    year, month = week_months[i]
    label = f"{month_names.get(str(month), month)} {year}"
    if start_col != end_col:
        ws.merge_cells(start_row=HEADER_ROW_MONTH, start_column=start_col, end_row=HEADER_ROW_MONTH, end_column=end_col)
    ws.cell(row=HEADER_ROW_MONTH, column=start_col, value=label)
    for cc in range(start_col, end_col + 1):
        mc = ws.cell(row=HEADER_ROW_MONTH, column=cc)
        mc.fill = PatternFill("solid", fgColor=S["month_fill"])
        mc.font = Font(name=FONT_NAME, size=11, bold=True, color=S["header_text"])
        mc.alignment = Alignment(horizontal="center", vertical="center")
        mc.border = border_all
    i = j + 1

ws.merge_cells(start_row=HEADER_ROW_MONTH, start_column=1, end_row=HEADER_ROW_WEEK, end_column=2)
hc = ws.cell(row=HEADER_ROW_MONTH, column=1, value=meta.get("epicColumnLabel", "Epic / Task"))
hc.fill = PatternFill("solid", fgColor=S["header_fill"])
hc.font = Font(name=FONT_NAME, size=11, bold=True, color=S["header_text"])
hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
hc.border = border_all
for cc in range(1, 3):
    ws.cell(row=HEADER_ROW_WEEK, column=cc).border = border_all
    ws.cell(row=HEADER_ROW_MONTH, column=cc).border = border_all

# ---------------- Week header (dd/mm over two lines) ----------------
for i, (wsd, wed) in enumerate(weeks):
    col = 3 + i
    label = f"{wsd.strftime('%d/%m')}\n{wed.strftime('%d/%m')}"
    cell = ws.cell(row=HEADER_ROW_WEEK, column=col, value=label)
    cell.fill = PatternFill("solid", fgColor=S["header_fill"])
    cell.font = Font(name=FONT_NAME, size=9, bold=True, color=S["header_text"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_all

ws.row_dimensions[HEADER_ROW_MONTH].height = 20
ws.row_dimensions[HEADER_ROW_WEEK].height = 28

# ---------------- Body: epic groups + task rows ----------------
row = FIRST_DATA_ROW
grouped = {epic: [t for t in tasks if t["epic"] == epic] for epic in epic_order}

for epic in epic_order:
    epic_tasks = grouped.get(epic, [])
    if not epic_tasks:
        continue
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + n_weeks)
    ec = ws.cell(row=row, column=1, value=epic)
    ec.font = Font(name=FONT_NAME, size=11, bold=True, color=S["epic_text"])
    ec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for cc in range(1, 3 + n_weeks):
        cell = ws.cell(row=row, column=cc)
        cell.border = border_all
        cell.fill = PatternFill("solid", fgColor=S["epic_fill"])
    ws.row_dimensions[row].height = 18
    row += 1

    for idx, t in enumerate(epic_tasks):
        t_start, t_end = pdate(t["start"]), pdate(t["end"])
        row_fill = S["alt_row"] if idx % 2 == 1 else "FFFFFF"

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        tc = ws.cell(row=row, column=1, value=t["title"])
        tc.font = Font(name=FONT_NAME, size=10, color=S["text"])
        tc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        tc.fill = PatternFill("solid", fgColor=row_fill)
        tc.border = border_all
        ws.cell(row=row, column=2).border = border_all
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=row_fill)

        for i, (wsd, wed) in enumerate(weeks):
            col = 3 + i
            cell = ws.cell(row=row, column=col)
            cell.border = border_all
            active = t_start <= wed and t_end >= wsd
            cell.fill = PatternFill("solid", fgColor=S["active_fill"] if active else row_fill)

        ws.row_dimensions[row].height = 16
        row += 1

# ---------------- Legend ----------------
row += 1
legend_cell = ws.cell(row=row, column=1, value=meta.get("legendPrefix", "Legend:"))
legend_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=S["text"])
sw = ws.cell(row=row, column=3)
sw.fill = PatternFill("solid", fgColor=S["active_fill"])
sw.border = border_all
lt = ws.cell(row=row, column=4, value=meta.get("legendLabel", "Week with an active task"))
lt.font = Font(name=FONT_NAME, size=10, color=S["muted"])

# ---------------- Page setup ----------------
ws.freeze_panes = f"C{FIRST_DATA_ROW}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.4
ws.print_area = f"A1:{get_column_letter(2 + n_weeks)}{row}"

wb.save(OUTPUT_PATH)
print(f"Wrote {OUTPUT_PATH} — style=yasdu, {row} rows, {len(tasks)} tasks, {n_weeks} weeks")

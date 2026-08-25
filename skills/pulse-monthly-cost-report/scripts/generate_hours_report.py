"""
Generate a 4-sheet monthly Hours & Cost report (.xlsx), always in the Yasdu
brand style, from a Pulse task list:
  1. Detalle <mes> — one row per task-resource pair, with business-day overlap
     against the target month, a 160-hs (or custom) monthly cap applied
     proportionally per resource, hourly rate lookup, and line cost.
  2. Costo por Recurso — one row per resource with a yellow input cell for
     monthly cost, and an auto-computed hourly rate (monthly cost / cap hours).
  3. Resumen por Recurso — hours (capped/uncapped) and cost totals per resource.
  4. Resumen por Epic y Tarea — hours and cost rolled up by epic, then by task.

Usage: python3 generate_hours_report.py <input.json> <output.xlsx> [skill_dir]

<skill_dir> is this skill's root folder, used to locate the assets/yasdu/
logo. Omit it and the script infers its own location, which usually works.

See ../references/input_schema.md for the full shape of input.json.
"""
import sys
import os
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage

if len(sys.argv) not in (3, 4):
    print("Usage: python3 generate_hours_report.py <input.json> <output.xlsx> [skill_dir]")
    sys.exit(1)

INPUT_PATH, OUTPUT_PATH = sys.argv[1], sys.argv[2]
SKILL_DIR = sys.argv[3] if len(sys.argv) == 4 else os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOGO_PATH = os.path.join(SKILL_DIR, "assets", "yasdu", "logo-horizontal-light.png")

with open(INPUT_PATH, encoding="utf-8") as f:
    data = json.load(f)

meta = data["meta"]
tasks_in = data["tasks"]  # [{epic, title, start, end, assignees:[{name, allocationPercent}]}]
MONTH_START = data["monthStart"]  # "YYYY-MM-01"
MONTH_END = data["monthEnd"]      # "YYYY-MM-DD" last day of month
CAP_HOURS = float(data.get("capHours", 160))

my, mm, _ = MONTH_START.split("-")
MONTH_START_Y, MONTH_START_M = int(my), int(mm)
ey, em, ed = MONTH_END.split("-")
MONTH_END_Y, MONTH_END_M, MONTH_END_D = int(ey), int(em), int(ed)

# Flatten to (epic, title, start, end, [(name, alloc), ...])
tasks = [(t["epic"], t["title"], t["start"], t["end"],
          [(a["name"], a["allocationPercent"]) for a in t.get("assignees", [])])
         for t in tasks_in]

# ---------------- Styling: always Yasdu brand ----------------
BLUE = "D85A28"       # Yasdu orange — primary header fill
DARK_BLUE = "123359"  # Yasdu navy — secondary header fill
ALT_ROW = "F4F2EC"
WHITE = "FFFFFF"
DARK = "1F2330"
GREY = "6E7180"
BORDER_GREY = "E2DFD9"
YELLOW = "FFF9C4"
FONT_NAME = meta.get("font", "Calibri")
ROW_OFFSET = 2  # rows reserved at the top of the Detalle sheet for the logo

thin = Side(style="thin", color=BORDER_GREY)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ==================== Sheet 1: Detalle ====================
ws = wb.active
detail_sheet_name = meta.get("detailSheetName", "Detalle")
ws.title = detail_sheet_name[:31]

if os.path.exists(LOGO_PATH):
    img = XLImage(LOGO_PATH)
    px_w, px_h = 1540, 443
    target_h_px = 28
    scale = target_h_px / px_h
    img.width = px_w * scale
    img.height = target_h_px
    ws.add_image(img, "A1")
    ws.row_dimensions[1].height = 24

headers = ["Epic", "Tarea", "Recurso", "Asignación %", "Inicio Tarea", "Fin Tarea",
           f"Inicio en {meta.get('monthLabel','Mes')}", f"Fin en {meta.get('monthLabel','Mes')}",
           "Días Hábiles en el Mes",
           "Horas en el Mes (sin tope)", "Horas Totales del Recurso (sin tope)",
           "Factor de Ajuste", f"Horas en el Mes (con tope de {int(CAP_HOURS)})",
           "Tarifa Horaria (USD)", "Costo (USD)"]
widths = [24, 32, 20, 13, 13, 13, 15, 13, 16, 16, 18, 14, 20, 16, 16]

HEADER_ROW = 1 + ROW_OFFSET
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    col = get_column_letter(i)
    ws.column_dimensions[col].width = w
    c = ws.cell(row=HEADER_ROW, column=i, value=h)
    c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_all
ws.row_dimensions[HEADER_ROW].height = 40
ws.freeze_panes = f"A{HEADER_ROW + 1}"

row = HEADER_ROW + 1
for epic, title, start, end, assignees in tasks:
    entries = assignees if assignees else [("— (sin asignar)", None)]
    for resource, alloc in entries:
        fill = ALT_ROW if (row % 2 == 0) else WHITE
        ws.cell(row=row, column=1, value=epic)
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value=resource)
        if alloc is not None:
            cell = ws.cell(row=row, column=4, value=alloc / 100)
            cell.number_format = "0%"
        e_cell = ws.cell(row=row, column=5, value=datetime.date.fromisoformat(start))
        e_cell.number_format = "yyyy-mm-dd"
        f_cell = ws.cell(row=row, column=6, value=datetime.date.fromisoformat(end))
        f_cell.number_format = "yyyy-mm-dd"

        g_cell = ws.cell(row=row, column=7, value=f"=MAX(E{row},DATE({MONTH_START_Y},{MONTH_START_M},1))")
        g_cell.number_format = "yyyy-mm-dd"
        h_cell = ws.cell(row=row, column=8, value=f"=MIN(F{row},DATE({MONTH_END_Y},{MONTH_END_M},{MONTH_END_D}))")
        h_cell.number_format = "yyyy-mm-dd"

        i_cell = ws.cell(row=row, column=9, value=f"=IF(H{row}<G{row},0,NETWORKDAYS(G{row},H{row}))")

        if alloc is not None:
            j_cell = ws.cell(row=row, column=10, value=f"=I{row}*8*D{row}")
        else:
            j_cell = ws.cell(row=row, column=10, value=0)
        j_cell.number_format = "0.0"

        k_cell = ws.cell(row=row, column=11, value=f'=SUMIFS($J$2:$J$5000,$C$2:$C$5000,C{row})')
        k_cell.number_format = "0.0"

        l_cell = ws.cell(row=row, column=12, value=f"=IF(K{row}=0,1,MIN(1,{CAP_HOURS}/K{row}))")
        l_cell.number_format = "0.00"

        m_cell = ws.cell(row=row, column=13, value=f"=J{row}*L{row}")
        m_cell.number_format = "0.0"

        n_cell = ws.cell(
            row=row, column=14,
            value=f"=IFERROR(INDEX(CostoPorRecurso[Tarifa Horaria (USD)],MATCH(C{row},CostoPorRecurso[Recurso],0)),0)"
        )
        n_cell.number_format = "$#,##0.00"

        o_cell = ws.cell(row=row, column=15, value=f"=M{row}*N{row}")
        o_cell.number_format = "$#,##0.00"

        for col in range(1, 16):
            cc = ws.cell(row=row, column=col)
            cc.font = Font(name=FONT_NAME, size=10, color=DARK)
            cc.fill = PatternFill("solid", fgColor=fill)
            cc.border = border_all
            if col >= 4:
                cc.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

LAST_ROW = row - 1

note_row = LAST_ROW + 2
note = ws.cell(
    row=note_row, column=1,
    value=(f"Supuesto: Horas en el mes (sin tope) = Días hábiles (lun-vie) dentro del solapamiento de la tarea con "
           f"{meta.get('monthLabel','el mes')} × 8 hs/día × % de asignación del recurso. No se aplicó calendario de feriados. "
           f"Tope: ningún recurso puede superar {int(CAP_HOURS)} hs/mes; si su total sin tope lo supera, el Factor de Ajuste "
           f"({int(CAP_HOURS)} / total sin tope) reduce proporcionalmente las horas de cada una de sus tareas hasta que el total "
           f"con tope sea {int(CAP_HOURS)}. Costo: Tarifa Horaria = Costo Mensual ingresado en la hoja 'Costo por Recurso' / "
           f"{int(CAP_HOURS)} hs. Costo de cada línea = Horas (con tope) × Tarifa Horaria. "
           f"Tareas sin recurso asignado en Pulse muestran 0 horas y 0 costo.")
)
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=15)
note.font = Font(name=FONT_NAME, size=9, italic=True, color=GREY)
note.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[note_row].height = 45

tab = Table(displayName="DetalleMes", ref=f"A{HEADER_ROW}:O{LAST_ROW}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=False)
ws.add_table(tab)

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.4
ws.print_area = f"A1:O{note_row}"

# Formula column-name refs use the Detalle table; fix N-column formula to
# use the actual table name chosen above (DetalleMes), and the resource
# total (K) self-reference to only sum this table's column, not raw ranges,
# for correctness when appended after other sheets. Re-point K formulas:
for r in range(HEADER_ROW + 1, LAST_ROW + 1):
    ws.cell(row=r, column=11).value = f'=SUMIFS(DetalleMes[Horas en el Mes (sin tope)],DetalleMes[Recurso],C{r})'

# ==================== Sheet 2: Costo por Recurso ====================
ws_cost = wb.create_sheet("Costo por Recurso")
resource_names = sorted({r for _, _, _, _, a in tasks for r, _ in a})

if os.path.exists(LOGO_PATH):
    img_cost = XLImage(LOGO_PATH)
    px_w, px_h = 1540, 443
    target_h_px = 28
    scale = target_h_px / px_h
    img_cost.width = px_w * scale
    img_cost.height = target_h_px
    ws_cost.add_image(img_cost, "A1")
    ws_cost.row_dimensions[1].height = 24

INSTR_ROW = 1 + ROW_OFFSET
instructions = ws_cost.cell(
    row=INSTR_ROW, column=1,
    value=(f"Completar el Costo Mensual (USD) de cada recurso en la columna resaltada en amarillo. "
           f"La Tarifa Horaria se calcula automáticamente como Costo Mensual / {int(CAP_HOURS)} hs (tope mensual de horas).")
)
ws_cost.merge_cells(start_row=INSTR_ROW, start_column=1, end_row=INSTR_ROW, end_column=3)
instructions.font = Font(name=FONT_NAME, size=9, italic=True, color=GREY)
instructions.alignment = Alignment(wrap_text=True, vertical="top")
ws_cost.row_dimensions[INSTR_ROW].height = 28

headers_cost = ["Recurso", "Costo Mensual (USD)", "Tarifa Horaria (USD)"]
widths_cost = [26, 20, 20]
HEADER_ROW_COST = INSTR_ROW + 2
for i, (h, w) in enumerate(zip(headers_cost, widths_cost), start=1):
    col = get_column_letter(i)
    ws_cost.column_dimensions[col].width = w
    c = ws_cost.cell(row=HEADER_ROW_COST, column=i, value=h)
    c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_all
ws_cost.row_dimensions[HEADER_ROW_COST].height = 26
ws_cost.freeze_panes = f"A{HEADER_ROW_COST + 1}"

rc = HEADER_ROW_COST + 1
for name in resource_names:
    fill = ALT_ROW if (rc % 2 == 0) else WHITE
    name_cell = ws_cost.cell(row=rc, column=1, value=name)
    name_cell.font = Font(name=FONT_NAME, size=10, color=DARK)
    name_cell.fill = PatternFill("solid", fgColor=fill)
    name_cell.border = border_all

    cost_cell = ws_cost.cell(row=rc, column=2, value=None)
    cost_cell.number_format = "$#,##0.00"
    cost_cell.fill = PatternFill("solid", fgColor=YELLOW)
    cost_cell.border = border_all
    cost_cell.alignment = Alignment(horizontal="center", vertical="center")
    cost_cell.comment = Comment("Ingresar el costo mensual (USD) de este recurso.", "Pulse Roadmap Toolkit")

    rate_cell = ws_cost.cell(row=rc, column=3, value=f'=IF(B{rc}="",0,B{rc}/{CAP_HOURS})')
    rate_cell.number_format = "$#,##0.00"
    rate_cell.fill = PatternFill("solid", fgColor=fill)
    rate_cell.border = border_all
    rate_cell.alignment = Alignment(horizontal="center", vertical="center")
    rate_cell.font = Font(name=FONT_NAME, size=10, color=DARK)
    rc += 1

LAST_COST_ROW = rc - 1
tab_cost = Table(displayName="CostoPorRecurso", ref=f"A{HEADER_ROW_COST}:C{LAST_COST_ROW}")
tab_cost.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=False)
ws_cost.add_table(tab_cost)

ws_cost.page_setup.fitToWidth = 1
ws_cost.page_setup.fitToHeight = 0
ws_cost.sheet_properties.pageSetUpPr.fitToPage = True
ws_cost.print_area = f"A1:C{LAST_COST_ROW}"

# ==================== Sheet 3: Resumen por Recurso ====================
ws2 = wb.create_sheet("Resumen por Recurso")

if os.path.exists(LOGO_PATH):
    img2 = XLImage(LOGO_PATH)
    px_w, px_h = 1540, 443
    target_h_px = 28
    scale = target_h_px / px_h
    img2.width = px_w * scale
    img2.height = target_h_px
    ws2.add_image(img2, "A1")
    ws2.row_dimensions[1].height = 24

headers2 = ["Recurso", f"Horas Totales (con tope de {int(CAP_HOURS)})", "Horas Totales (sin tope)",
            "¿Tope aplicado?", "Cantidad de Tareas", "Costo Total (USD)"]
widths2 = [26, 22, 20, 16, 18, 18]
HEADER_ROW2 = 1 + ROW_OFFSET
for i, (h, w) in enumerate(zip(headers2, widths2), start=1):
    col = get_column_letter(i)
    ws2.column_dimensions[col].width = w
    c = ws2.cell(row=HEADER_ROW2, column=i, value=h)
    c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_all
ws2.row_dimensions[HEADER_ROW2].height = 26
ws2.freeze_panes = f"A{HEADER_ROW2 + 1}"

r2 = HEADER_ROW2 + 1
FIRST_ROW2 = r2
for name in resource_names:
    fill = ALT_ROW if (r2 % 2 == 0) else WHITE
    ws2.cell(row=r2, column=1, value=name)
    hcell = ws2.cell(row=r2, column=2, value=f'=SUMIFS(DetalleMes[Horas en el Mes (con tope de {int(CAP_HOURS)})],DetalleMes[Recurso],A{r2})')
    hcell.number_format = "0.0"
    ucell = ws2.cell(row=r2, column=3, value=f'=SUMIFS(DetalleMes[Horas en el Mes (sin tope)],DetalleMes[Recurso],A{r2})')
    ucell.number_format = "0.0"
    fcell = ws2.cell(row=r2, column=4, value=f'=IF(C{r2}>{CAP_HOURS},"Sí","No")')
    ccell = ws2.cell(row=r2, column=5, value=f'=COUNTIFS(DetalleMes[Recurso],A{r2},DetalleMes[Horas en el Mes (con tope de {int(CAP_HOURS)})],">0")')
    costcell = ws2.cell(row=r2, column=6, value=f'=SUMIFS(DetalleMes[Costo (USD)],DetalleMes[Recurso],A{r2})')
    costcell.number_format = "$#,##0.00"
    for col in range(1, 7):
        cc = ws2.cell(row=r2, column=col)
        cc.font = Font(name=FONT_NAME, size=10, color=DARK)
        cc.fill = PatternFill("solid", fgColor=fill)
        cc.border = border_all
        if col >= 2:
            cc.alignment = Alignment(horizontal="center", vertical="center")
    r2 += 1

LAST_ROW2 = r2 - 1
tab2 = Table(displayName="ResumenRecursos", ref=f"A{HEADER_ROW2}:F{LAST_ROW2}")
tab2.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=False)
ws2.add_table(tab2)

tot_row = LAST_ROW2 + 1
ws2.cell(row=tot_row, column=1, value="Total").font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
tcell = ws2.cell(row=tot_row, column=2, value=f"=SUM(B{FIRST_ROW2}:B{LAST_ROW2})")
tcell.number_format = "0.0"
tcell.font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
utcell = ws2.cell(row=tot_row, column=3, value=f"=SUM(C{FIRST_ROW2}:C{LAST_ROW2})")
utcell.number_format = "0.0"
utcell.font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
costtot = ws2.cell(row=tot_row, column=6, value=f"=SUM(F{FIRST_ROW2}:F{LAST_ROW2})")
costtot.number_format = "$#,##0.00"
costtot.font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
for col in range(1, 7):
    ws2.cell(row=tot_row, column=col).border = border_all

ws2.page_setup.orientation = "landscape"
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.sheet_properties.pageSetUpPr.fitToPage = True
ws2.page_margins.left = 0.3
ws2.page_margins.right = 0.3
ws2.print_area = f"A1:F{tot_row}"

# ==================== Sheet 4: Resumen por Epic y Tarea ====================
ws3 = wb.create_sheet("Resumen por Epic y Tarea")

if os.path.exists(LOGO_PATH):
    img3 = XLImage(LOGO_PATH)
    px_w, px_h = 1540, 443
    target_h_px = 28
    scale = target_h_px / px_h
    img3.width = px_w * scale
    img3.height = target_h_px
    ws3.add_image(img3, "A1")
    ws3.row_dimensions[1].height = 24

epic_order, seen_epics = [], set()
task_order, seen_tasks = [], set()
for epic, title, *_ in tasks:
    if epic not in seen_epics:
        seen_epics.add(epic)
        epic_order.append(epic)
    key = (epic, title)
    if key not in seen_tasks:
        seen_tasks.add(key)
        task_order.append(key)

TITLE_A_ROW = 1 + ROW_OFFSET
title_a = ws3.cell(row=TITLE_A_ROW, column=1, value="Horas y Costo por Epic")
title_a.font = Font(name=FONT_NAME, size=12, bold=True, color=BLUE)
ws3.merge_cells(start_row=TITLE_A_ROW, start_column=1, end_row=TITLE_A_ROW, end_column=3)

headers_a = ["Epic", f"Horas Totales (con tope de {int(CAP_HOURS)})", "Costo Total (USD)"]
widths_a = [32, 26, 18]
HEADER_ROW_A = TITLE_A_ROW + 2
for i, (h, w) in enumerate(zip(headers_a, widths_a), start=1):
    col = get_column_letter(i)
    ws3.column_dimensions[col].width = w
    c = ws3.cell(row=HEADER_ROW_A, column=i, value=h)
    c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_all
ws3.row_dimensions[HEADER_ROW_A].height = 26

ra = HEADER_ROW_A + 1
FIRST_A = ra
for epic in epic_order:
    fill = ALT_ROW if (ra % 2 == 0) else WHITE
    ws3.cell(row=ra, column=1, value=epic)
    hcell = ws3.cell(row=ra, column=2, value=f'=SUMIFS(DetalleMes[Horas en el Mes (con tope de {int(CAP_HOURS)})],DetalleMes[Epic],A{ra})')
    hcell.number_format = "0.0"
    ccell = ws3.cell(row=ra, column=3, value=f'=SUMIFS(DetalleMes[Costo (USD)],DetalleMes[Epic],A{ra})')
    ccell.number_format = "$#,##0.00"
    for col in range(1, 4):
        cc = ws3.cell(row=ra, column=col)
        cc.font = Font(name=FONT_NAME, size=10, color=DARK)
        cc.fill = PatternFill("solid", fgColor=fill)
        cc.border = border_all
        if col >= 2:
            cc.alignment = Alignment(horizontal="center", vertical="center")
    ra += 1

LAST_A = ra - 1
tab_a = Table(displayName="ResumenEpic", ref=f"A{HEADER_ROW_A}:C{LAST_A}")
tab_a.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=False)
ws3.add_table(tab_a)

tot_a = LAST_A + 1
ws3.cell(row=tot_a, column=1, value="Total").font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
tcell_a = ws3.cell(row=tot_a, column=2, value=f"=SUM(B{FIRST_A}:B{LAST_A})")
tcell_a.number_format = "0.0"
tcell_a.font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
ccell_a = ws3.cell(row=tot_a, column=3, value=f"=SUM(C{FIRST_A}:C{LAST_A})")
ccell_a.number_format = "$#,##0.00"
ccell_a.font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
for col in (1, 2, 3):
    ws3.cell(row=tot_a, column=col).border = border_all

title_b_row = tot_a + 3
title_b = ws3.cell(row=title_b_row, column=1, value="Horas y Costo por Tarea")
title_b.font = Font(name=FONT_NAME, size=12, bold=True, color=BLUE)
ws3.merge_cells(start_row=title_b_row, start_column=1, end_row=title_b_row, end_column=4)

headers_b = ["Epic", "Tarea", f"Horas Totales (con tope de {int(CAP_HOURS)})", "Costo Total (USD)"]
widths_b = [24, 34, 26, 18]
HEADER_ROW_B = title_b_row + 2
for i, (h, w) in enumerate(zip(headers_b, widths_b), start=1):
    col = get_column_letter(i)
    ws3.column_dimensions[col].width = max(ws3.column_dimensions[col].width or 0, w)
    c = ws3.cell(row=HEADER_ROW_B, column=i, value=h)
    c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_all
ws3.row_dimensions[HEADER_ROW_B].height = 26

rb = HEADER_ROW_B + 1
FIRST_B = rb
for epic, title in task_order:
    fill = ALT_ROW if (rb % 2 == 0) else WHITE
    ws3.cell(row=rb, column=1, value=epic)
    ws3.cell(row=rb, column=2, value=title)
    hcell = ws3.cell(
        row=rb, column=3,
        value=f'=SUMIFS(DetalleMes[Horas en el Mes (con tope de {int(CAP_HOURS)})],DetalleMes[Tarea],B{rb},DetalleMes[Epic],A{rb})'
    )
    hcell.number_format = "0.0"
    ccell = ws3.cell(
        row=rb, column=4,
        value=f'=SUMIFS(DetalleMes[Costo (USD)],DetalleMes[Tarea],B{rb},DetalleMes[Epic],A{rb})'
    )
    ccell.number_format = "$#,##0.00"
    for col in range(1, 5):
        cc = ws3.cell(row=rb, column=col)
        cc.font = Font(name=FONT_NAME, size=10, color=DARK)
        cc.fill = PatternFill("solid", fgColor=fill)
        cc.border = border_all
        if col >= 3:
            cc.alignment = Alignment(horizontal="center", vertical="center")
    rb += 1

LAST_B = rb - 1
tab_b = Table(displayName="ResumenTarea", ref=f"A{HEADER_ROW_B}:D{LAST_B}")
tab_b.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=False)
ws3.add_table(tab_b)

tot_b = LAST_B + 1
ws3.cell(row=tot_b, column=2, value="Total").font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
tcell_b = ws3.cell(row=tot_b, column=3, value=f"=SUM(C{FIRST_B}:C{LAST_B})")
tcell_b.number_format = "0.0"
tcell_b.font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
ccell_b = ws3.cell(row=tot_b, column=4, value=f"=SUM(D{FIRST_B}:D{LAST_B})")
ccell_b.number_format = "$#,##0.00"
ccell_b.font = Font(name=FONT_NAME, size=10, bold=True, color=DARK)
for col in (1, 2, 3, 4):
    ws3.cell(row=tot_b, column=col).border = border_all

ws3.page_setup.orientation = "landscape"
ws3.page_setup.fitToWidth = 1
ws3.page_setup.fitToHeight = 0
ws3.sheet_properties.pageSetUpPr.fitToPage = True
ws3.page_margins.left = 0.3
ws3.page_margins.right = 0.3
ws3.print_area = f"A1:D{tot_b}"

wb.save(OUTPUT_PATH)
print(f"Wrote {OUTPUT_PATH} — {len(tasks)} tasks, {LAST_ROW-1} detail rows, {len(resource_names)} resources, cap={CAP_HOURS}")
